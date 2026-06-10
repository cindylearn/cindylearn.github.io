#!/usr/bin/env python3
"""
SUMA Ads Dashboard — Weekly Auto-Updater
Runs every Monday at 12pm via launchd.

每周一 12pm 自动执行：
  1. 下载 FY26 Google Sheet
  2. 读取每个渠道的月份数据 → 更新 HTML 里的 RAW 数组
  3. 从 总览 tab 更新 VERIFIED.FY26 年度汇总
  4. 重新计算 Q2 OKR actuals（Apr+May+Jun 的 QL/Interview/Students）
  5. Git commit + push → GitHub Pages 自动部署

Usage:
  python3 update_dashboard.py              # 正常运行（更新 + push）
  python3 update_dashboard.py --dry-run    # 更新 HTML 但不 push
  python3 update_dashboard.py --status     # 只显示当前数据，不修改任何东西
"""

import re
import io
import shutil
import subprocess
import sys
import traceback
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

FY26_SHEET_ID = "1JHulpfu2gEkBXadX-tZJYxBEr2ffru3TguVqbjA6J9E"  # Oct 2025 – ongoing
# FY25_SHEET_ID = "1NCxLbz8zvPYz2Ezh17zz9XEa0neMintnPnDYbJ_soX0"  # Historical, not changed

SHEET_URL = "https://docs.google.com/spreadsheets/d/{id}/export?format=xlsx"

# Month index: Oct 2024 = 0, Nov 2024 = 1, ..., Oct 2025 = 12, Apr 2026 = 18 ...
BASE_YEAR, BASE_MONTH = 2024, 10

# Dashboard ML array — must stay in sync with HTML
ML = [
    'Oct 24','Nov 24','Dec 24','Jan 25','Feb 25','Mar 25',
    'Apr 25','May 25','Jun 25','Jul 25','Aug 25','Sep 25',
    'Oct 25','Nov 25','Dec 25','Jan 26','Feb 26','Mar 26',
    'Apr 26','May 26',
    # 'Jun 26',   ← uncomment + add to HTML when June data arrives
]

# Q2 Apr–Jun 2026
Q2_MONTHS = {(2026, 4), (2026, 5), (2026, 6)}

# Channels: FY26 tab name fragment → dashboard channel name
CHANNEL_TABS = {
    '01 Lead Magnet':  'Lead Magnet',
    '02 VSL Ads':      'VSL',
    '03 Trial Class':  'Trial Class',
    '04 Open Day':     'Open Day',
    '05 Google Ads':   'Google Ads',
    '06 Youtube Ads':  'YouTube',
    '07 Road Show':    'Road Show',
    '08 School Fair':  'School Fair',
    '09 XHS':          'XHS',
    '10 KOL':          'KOL Influencer',
}

DIGITAL_CH = ['VSL', 'Trial Class', 'Open Day', 'Google Ads', 'XHS', 'YouTube', 'Lead Magnet']
ALL_CH = DIGITAL_CH + ['KOL Influencer', 'Road Show', 'School Fair']

# 总览 tab: exact first-line channel label → dashboard channel name
# Order matters: longer/more-specific strings must come first
ZONGLAN_MAP = [
    ('Google Ads',  'Google Ads'),
    ('Youtube Ads', 'YouTube'),
    ('Youtube',     'YouTube'),
    ('Lead Magnet', 'Lead Magnet'),
    ('Trial Class', 'Trial Class'),
    ('Open Day',    'Open Day'),
    ('School Fair', 'School Fair'),
    ('Road Show',   'Road Show'),
    ('KOL',         'KOL Influencer'),
    ('XHS',         'XHS'),
    ('Ads',         'VSL'),         # Must be LAST — "Ads" matches the VSL row only
]

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────

DASHBOARD_SRC = Path("/Users/suma/Downloads/SUMA_Ads_Dashboard.html")
REPO_DIR      = Path("/Users/suma/Downloads/cindylearn-pages")
REPO_HTML     = REPO_DIR / "tools-internal/suma-ads-dashboard/index.html"
LOG_FILE      = REPO_DIR / "update.log"

# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def log(msg):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def month_to_idx(year: int, month: int) -> int:
    """Oct 2024 = 0, Nov 2024 = 1, ..., Oct 2025 = 12, Apr 2026 = 18 ..."""
    return (year - BASE_YEAR) * 12 + (month - BASE_MONTH)


def idx_to_month(idx: int) -> tuple:
    """Inverse of month_to_idx. Returns (year, month)."""
    total_months = idx + BASE_MONTH - 1
    year  = BASE_YEAR + total_months // 12
    month = total_months % 12 + 1
    return year, month


def num(v, default=0.0) -> float:
    """Convert cell value to float."""
    try:
        return float(v) if v is not None and str(v).strip() not in ('', '#DIV/0!') else default
    except (TypeError, ValueError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Download sheet
# ─────────────────────────────────────────────────────────────────────────────

def download_sheet(sheet_id: str) -> bytes:
    import browser_cookie3
    import requests as rq
    url     = SHEET_URL.format(id=sheet_id)
    cookies = browser_cookie3.chrome(domain_name=".google.com")
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    r = rq.get(url, cookies=cookies, headers=headers, timeout=30)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} for sheet {sheet_id}")
    return r.content


def open_wb(data: bytes):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Parse channel tabs → monthly data per channel
# ─────────────────────────────────────────────────────────────────────────────

def parse_channel_tabs(wb) -> dict:
    """
    Returns:
      { channel_name: { idx: { spend, leads, quality, signup, interview, students } } }
    where idx = dashboard month index (Oct 2024 = 0).
    """
    result = {}

    for frag, ch_name in CHANNEL_TABS.items():
        tab = next((t for t in wb.sheetnames if frag.lower() in t.lower()), None)
        if not tab:
            log(f"  WARNING: tab '{frag}' not found in FY26 sheet")
            continue

        ws   = wb[tab]
        data = {}

        for row in range(3, ws.max_row + 1):
            raw_date = ws.cell(row=row, column=1).value
            if raw_date is None:
                continue
            try:
                if hasattr(raw_date, "month"):
                    year, month = raw_date.year, raw_date.month
                else:
                    dt    = datetime.strptime(str(raw_date).split()[0], "%Y-%m-%d")
                    year, month = dt.year, dt.month
            except Exception:
                continue

            idx = month_to_idx(year, month)
            # Only include months present in the ML array
            if idx < 0 or idx >= len(ML):
                continue

            data[idx] = {
                "spend":     num(ws.cell(row=row, column=3).value),
                "leads":     int(num(ws.cell(row=row, column=4).value)),
                "quality":   int(num(ws.cell(row=row, column=5).value)),
                "signup":    int(num(ws.cell(row=row, column=6).value)),
                "interview": int(num(ws.cell(row=row, column=7).value)),
                "students":  int(num(ws.cell(row=row, column=8).value)),
            }

        if data:
            result[ch_name] = data
            latest_idx = max(data.keys())
            ly, lm = idx_to_month(latest_idx)
            log(f"  {ch_name:18s} → {len(data)} months (latest: {lm}/{ly})")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Parse 总览 tab → VERIFIED.FY26
# ─────────────────────────────────────────────────────────────────────────────

def parse_verified_fy26(wb) -> dict:
    """
    Read '总览' tab → find rows with '2026 年度' → return VERIFIED.FY26 dict.
    Returns: { channel: { spend, leads, quality, signup, interview, students } }
    """
    ws = wb['总览']
    result   = {}
    cur_ch   = None

    for row in range(1, ws.max_row + 1):
        col_a = str(ws.cell(row=row, column=1).value or '').strip()
        col_b = str(ws.cell(row=row, column=2).value or '').strip()

        # Detect channel name in col A
        if col_a and col_a not in ('', 'Total', '总计', 'Grand Total'):
            # Normalise: take first line, strip whitespace
            col_a_first = col_a.split('\n')[0].strip()
            matched = None
            for zl_key, dash in ZONGLAN_MAP:
                # Exact match OR starts-with match (handles "Ads" → row "Ads" only)
                if col_a_first.lower() == zl_key.lower() or col_a_first.lower().startswith(zl_key.lower() + ' '):
                    matched = dash
                    break
            if matched:
                cur_ch = matched

        # Detect '2026 年度' row (note: sheet uses '2026 年度' with a space)
        if cur_ch and ('2026' in col_b and '年度' in col_b):
            c = lambda col: num(ws.cell(row=row, column=col).value)
            result[cur_ch] = {
                "spend":     round(c(3), 2),
                "leads":     int(c(4)),
                "quality":   int(c(5)),
                "signup":    int(c(6)),
                "interview": int(c(7)),
                "students":  int(c(8)),
            }
            log(f"  VERIFIED {cur_ch:18s} → spend={result[cur_ch]['spend']:.2f}  students={result[cur_ch]['students']}")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Compute Q2 actuals from channel data
# ─────────────────────────────────────────────────────────────────────────────

def compute_q2_actuals(ch_data: dict) -> dict:
    """Sum quality/interview/students across Q2 months for all DIGITAL_CH."""
    q2_idx = {month_to_idx(y, m) for y, m in Q2_MONTHS}
    totals  = {"ql": 0, "int": 0, "paid": 0}

    for ch in DIGITAL_CH:
        for idx, d in ch_data.get(ch, {}).items():
            if idx in q2_idx:
                totals["ql"]   += d["quality"]
                totals["int"]  += d["interview"]
                totals["paid"] += d["students"]

    return totals


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Patch HTML
# ─────────────────────────────────────────────────────────────────────────────

def extract_raw_block(html: str) -> tuple:
    """Return (start_idx, end_idx) of the const RAW = { ... } block."""
    start = html.find("const RAW = {")
    if start == -1:
        raise RuntimeError("'const RAW = {' not found")
    depth = 0
    for i in range(start, len(html)):
        if html[i] == "{":
            depth += 1
        elif html[i] == "}":
            depth -= 1
            if depth == 0:
                return start, i + 1
    raise RuntimeError("Unclosed RAW block")


def parse_channel_arrays(raw_js: str, channel: str) -> dict:
    """Parse one channel's 6 arrays from RAW JS. Returns {metric: [floats]}."""
    ch_m = re.search(
        r"['\"]" + re.escape(channel) + r"['\"]" + r"\s*:\s*\{([^}]+(?:\{[^}]*\}[^}]*)*)\}",
        raw_js, re.DOTALL
    )
    if not ch_m:
        return {}
    block  = ch_m.group(1)
    arrays = {}
    for metric in ("spend", "leads", "quality", "signup", "interview", "students"):
        m = re.search(rf"{metric}\s*:\s*\[([^\]]+)\]", block)
        if m:
            nums = []
            for x in m.group(1).split(","):
                try:
                    nums.append(float(x.strip()))
                except ValueError:
                    nums.append(0.0)
            arrays[metric] = nums
    return arrays


def build_raw_block(raw_js: str, ch_data: dict) -> str:
    """
    Rebuild RAW JS block with updated FY26 data.
    For each channel, replaces indices 12+ with values from ch_data.
    Leaves FY25 data (indices 0-11) untouched.
    """
    new_raw_js = raw_js

    for ch in ALL_CH:
        arrays = parse_channel_arrays(raw_js, ch)
        if not arrays:
            continue

        sheet_months = ch_data.get(ch, {})  # { idx: {metrics} }

        # Update each metric array
        for metric in ("spend", "leads", "quality", "signup", "interview", "students"):
            nums = arrays.get(metric, [])
            # Ensure array is long enough
            while len(nums) < len(ML):
                nums.append(0.0)
            # Apply FY26 updates (indices 12+)
            for idx, d in sheet_months.items():
                if 12 <= idx < len(nums):
                    nums[idx] = d[metric]

            # Format array: spend with 2dp, rest as ints
            if metric == "spend":
                parts = [f"{v:.2f}" if v != int(v) else str(int(v)) for v in nums]
            else:
                parts = [str(int(v)) for v in nums]

            # Reconstruct the array string with spacing
            new_arr   = ", ".join(parts)
            old_pat   = re.compile(
                re.escape(metric) + r"\s*:\s*\[([^\]]+)\]"
            )
            # Only replace inside the channel block
            ch_start  = new_raw_js.find(f"'{ch}':")
            if ch_start == -1:
                ch_start = new_raw_js.find(f'"{ch}":')
            if ch_start == -1:
                continue
            # Work within the channel block (next ~3000 chars)
            window    = new_raw_js[ch_start : ch_start + 3000]
            new_window = old_pat.sub(f"{metric}:  [{new_arr}]", window, count=1)
            new_raw_js = new_raw_js[:ch_start] + new_window + new_raw_js[ch_start + 3000:]

    return new_raw_js


def patch_verified_fy26(html: str, fy26: dict) -> str:
    """Replace FY26: { ... } block in VERIFIED."""
    start = html.find("FY26: {")
    if start == -1:
        log("  WARN: 'FY26: {' not found")
        return html
    depth = 0
    for i in range(start + 7, len(html)):
        if html[i] == "{":   depth += 1
        elif html[i] == "}":
            if depth == 0: end = i + 1; break
            depth -= 1

    digital = ['VSL','Trial Class','Open Day','Google Ads','XHS','YouTube','Lead Magnet']
    offline = ['KOL Influencer','Road Show','School Fair']
    lines   = ["\n"]
    for ch in digital:
        d   = fy26.get(ch, {'spend':0,'leads':0,'quality':0,'signup':0,'interview':0,'students':0})
        pad = max(1, 14 - len(ch))
        lines.append(f"      '{ch}':{' '*pad}"
                     f"{{spend:{d['spend']}, leads:{d['leads']}, quality:{d['quality']}, "
                     f"signup:{d['signup']}, interview:{d['interview']}, students:{d['students']}}},")
    lines.append("      // Offline")
    for ch in offline:
        d   = fy26.get(ch, {'spend':0,'leads':0,'quality':0,'signup':0,'interview':0,'students':0})
        pad = max(1, 14 - len(ch))
        lines.append(f"      '{ch}':{' '*pad}"
                     f"{{spend:{d['spend']}, leads:{d['leads']}, quality:{d['quality']}, "
                     f"signup:{d['signup']}, interview:{d['interview']}, students:{d['students']}}},")
    lines.append("    ")
    return html[:start] + "FY26: {" + "\n".join(lines) + "}" + html[end:]


def patch_q2_actuals(html: str, actuals: dict) -> str:
    """Update actuals:{ ql, int, paid } in QUARTERS['26Q2']."""
    ql, intn, paid = actuals["ql"], actuals["int"], actuals["paid"]
    pos = html.find("'26Q2':")
    if pos == -1: pos = html.find('"26Q2":')
    if pos == -1:
        log("  WARN: '26Q2' block not found")
        return html
    window  = html[pos : pos + 2000]
    old_pat = re.compile(r'actuals:\s*\{\s*ql:\d+,\s*int:\d+,\s*paid:\d+\s*\}')
    m       = old_pat.search(window)
    if not m:
        log("  WARN: actuals pattern not found")
        return html
    new_win = window[:m.start()] + f"actuals:{{ ql:{ql}, int:{intn}, paid:{paid} }}" + window[m.end():]
    return html[:pos] + new_win + html[pos + 2000:]


def patch_raw(html: str, ch_data: dict) -> str:
    """Replace the RAW block with updated channel arrays."""
    start, end = extract_raw_block(html)
    old_raw    = html[start:end]
    new_raw    = build_raw_block(old_raw, ch_data)
    return html[:start] + new_raw + html[end:]


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Git push
# ─────────────────────────────────────────────────────────────────────────────

def git_push():
    today = datetime.now().strftime("%b %d, %Y")
    for cmd in [
        ["git", "-C", str(REPO_DIR), "add", "index.html"],
        ["git", "-C", str(REPO_DIR), "commit", "-m", f"Auto-update: {today}"],
        ["git", "-C", str(REPO_DIR), "push",   "origin", "main"],
    ]:
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            raise RuntimeError(f"git {' '.join(cmd[2:])}: {r.stderr.strip()}")
        log(f"  ✅ git {' '.join(cmd[2:])}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    args    = sys.argv[1:]
    dry_run = "--dry-run" in args
    status  = "--status"  in args

    # ── Status mode ──────────────────────────────────────────────────────────
    if status:
        log("Downloading FY26 sheet for status check…")
        data    = download_sheet(FY26_SHEET_ID)
        wb      = open_wb(data)
        ch_data = parse_channel_tabs(wb)
        actuals = compute_q2_actuals(ch_data)
        print(f"\nQ2 actuals (Apr+May+Jun from FY26 sheet):")
        print(f"  Quality Leads : {actuals['ql']}")
        print(f"  Interviews    : {actuals['int']}")
        print(f"  Paid/Enrolled : {actuals['paid']}")
        print(f"\nFY26 YTD per channel (from sheet tabs):")
        for ch in DIGITAL_CH:
            months = ch_data.get(ch, {})
            spend  = sum(d['spend']    for d in months.values())
            ql     = sum(d['quality']  for d in months.values())
            stu    = sum(d['students'] for d in months.values())
            print(f"  {ch:18s} spend={spend:>10.2f}  QL={ql:>4}  students={stu}")
        return

    log("=" * 60)
    log("SUMA Dashboard auto-update starting")

    # ── Download FY26 sheet ───────────────────────────────────────────────────
    try:
        log(f"Downloading FY26 sheet…")
        data    = download_sheet(FY26_SHEET_ID)
        wb      = open_wb(data)
        log(f"  Tabs: {len(wb.sheetnames)} found")
    except Exception as e:
        log(f"  FATAL: Cannot download FY26 sheet — {e}")
        log(traceback.format_exc())
        return

    # ── Parse channel tabs ────────────────────────────────────────────────────
    try:
        ch_data = parse_channel_tabs(wb)
        log(f"  Parsed {len(ch_data)} channels")
    except Exception as e:
        log(f"  ERROR parsing channel tabs: {e}")
        log(traceback.format_exc())
        ch_data = {}

    # ── Parse 总览 → VERIFIED.FY26 ────────────────────────────────────────────
    try:
        fy26_verified = parse_verified_fy26(wb)
        log(f"  Parsed VERIFIED.FY26 for {len(fy26_verified)} channels")
    except Exception as e:
        log(f"  ERROR parsing 总览: {e}")
        log(traceback.format_exc())
        fy26_verified = {}

    # ── Compute Q2 actuals ────────────────────────────────────────────────────
    actuals = compute_q2_actuals(ch_data)
    log(f"  Q2 actuals: ql={actuals['ql']}, int={actuals['int']}, paid={actuals['paid']}")

    # ── Patch HTML ────────────────────────────────────────────────────────────
    html    = DASHBOARD_SRC.read_text(encoding="utf-8")
    changed = False

    # 1. Update RAW arrays
    if ch_data:
        try:
            new_html = patch_raw(html, ch_data)
            if new_html != html:
                html    = new_html
                changed = True
                log("  ✅ RAW arrays updated")
            else:
                log("  RAW arrays unchanged")
        except Exception as e:
            log(f"  ERROR patching RAW: {e}")
            log(traceback.format_exc())

    # 2. Update VERIFIED.FY26
    if fy26_verified:
        try:
            new_html = patch_verified_fy26(html, fy26_verified)
            if new_html != html:
                html    = new_html
                changed = True
                log("  ✅ VERIFIED.FY26 updated")
            else:
                log("  VERIFIED.FY26 unchanged")
        except Exception as e:
            log(f"  ERROR patching VERIFIED: {e}")
            log(traceback.format_exc())

    # 3. Update Q2 actuals
    try:
        new_html = patch_q2_actuals(html, actuals)
        if new_html != html:
            html    = new_html
            changed = True
            log(f"  ✅ Q2 actuals → ql:{actuals['ql']}, int:{actuals['int']}, paid:{actuals['paid']}")
        else:
            log(f"  Q2 actuals unchanged")
    except Exception as e:
        log(f"  ERROR patching actuals: {e}")
        log(traceback.format_exc())

    # ── Save ──────────────────────────────────────────────────────────────────
    if changed:
        DASHBOARD_SRC.write_text(html, encoding="utf-8")
        shutil.copy2(DASHBOARD_SRC, REPO_HTML)
        log(f"  ✅ Saved HTML + copied to repo")
    else:
        log("  No changes to dashboard")

    # ── Git push ──────────────────────────────────────────────────────────────
    if dry_run:
        log("Dry-run mode — skipping git push")
    elif changed:
        try:
            git_push()
            log("✅ Live at https://cindylearn.github.io/tools-internal/suma-ads-dashboard/")
        except Exception as e:
            log(f"  ERROR git push: {e}")
    else:
        log("Nothing to push (no changes)")

    log("Done.")
    log("=" * 60)


if __name__ == "__main__":
    main()
